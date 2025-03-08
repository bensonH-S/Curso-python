import openpyxl
from fuzzywuzzy import fuzz

# Carregar as planilhas
wb1 = openpyxl.load_workbook(r'C:\Users\u512228\Documents\ATMs_P1.xlsx')
wb2 = openpyxl.load_workbook(r'C:\Users\u512228\Documents\ATM_P2.XLSX')

# Acessar as abas
ws1 = wb1['ATMe DADOS']
ws2 = wb2['ATM']

# Obter as chaves primárias (PK) de ambas as planilhas
pk_column1 = ws1['A']  # Coluna de PK da Planilha 1 (ajuste conforme necessário)
pk_column2 = ws2['A']  # Coluna de PK da Planilha 2 (ajuste conforme necessário)

# Obter os endereços para comparação
enderecos_planilha1 = [ws1[f'G{i}'].value for i in range(2, ws1.max_row + 1)]  # Ajuste para a coluna do endereço
enderecos_planilha2 = [ws2[f'K{i}'].value for i in range(2, ws2.max_row + 1)]  # Ajuste para a coluna do endereço

# Função para comparar endereços usando a similaridade
def comparar_enderecos(endereco1, endereco2):
    return fuzz.partial_ratio(endereco1, endereco2) > 50  # Ajuste a similaridade para 50%

# Comparar endereços das duas planilhas
enderecos_semelhantes = []

for pk1, endereco1 in zip(pk_column1[1:], enderecos_planilha1):  # Ignorar o cabeçalho (primeira linha)
    for pk2, endereco2 in zip(pk_column2[1:], enderecos_planilha2):  # Ignorar o cabeçalho (primeira linha)
        if endereco1 and endereco2:  # Ignorar células vazias
            if comparar_enderecos(endereco1, endereco2):  # Se os endereços são semelhantes
                enderecos_semelhantes.append((pk1.value, endereco1, pk2.value, endereco2))

# Gerar o arquivo de saída com as informações
if enderecos_semelhantes:
    with open(r'C:\Users\u512228\Documents\enderecos_semelhantes.txt', 'w', encoding='utf-8') as f:
        f.write("Endereços semelhantes encontrados:\n\n")
        for pk1, endereco1, pk2, endereco2 in enderecos_semelhantes:
            f.write(f"PK1: {pk1} - Endereço: {endereco1}\n")
            f.write(f"PK2: {pk2} - Endereço: {endereco2}\n")
            f.write("-" * 40 + "\n")
    print("Arquivo de endereços semelhantes gerado com sucesso.")
else:
    with open(r'C:\Users\u512228\Documents\enderecos_semelhantes.txt', 'w', encoding='utf-8') as f:
        f.write("Nenhum endereço semelhante encontrado.\n")
    print("Nenhum endereço semelhante encontrado. Arquivo gerado.")
