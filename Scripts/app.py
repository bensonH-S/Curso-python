import streamlit as st
import pandas as pd
from sqlalchemy import create_engine
import logging
import openpyxl
from datetime import datetime

# Configuração da página para modo wide (expandido)
st.set_page_config(page_title="Dashboard GECAF", layout="wide")

# Configuração do log
logging.basicConfig(level=logging.INFO)

# Conexão com o banco de dados usando SQLAlchemy
DATABASE_URI = 'mariadb+mariadbconnector://root:@localhost:3306/gecaf'
engine = create_engine(DATABASE_URI)

def tratar_valores(valor):
    if valor is None:
        return None
    if isinstance(valor, str):
        return valor.strip()
    if isinstance(valor, (int, float)):
        return valor
    if isinstance(valor, bool):
        return int(valor)
    if isinstance(valor, datetime):
        return valor.strftime('%Y-%m-%d')
    return str(valor)

def extrair_dados(planilha):
    wb = openpyxl.load_workbook(planilha, data_only=True)
    dados_107 = []
    dados_005 = []
    for aba in ["107", "005"]:
        if aba in wb.sheetnames:
            sheet = wb[aba]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None and row[0] != "" and row[0] != "registro":
                    linha_formatada = tuple(tratar_valores(celula) for celula in row)
                    if aba == "107":
                        dados_107.append(linha_formatada)
                    else:
                        dados_005.append(linha_formatada)
                    # logging.info(f"Dados formatados para {aba}: {linha_formatada}")
    return dados_107, dados_005

def inserir_dados(tabela, dados):
    if not dados:
        return
    try:
        conn = engine.raw_connection()
        cursor = conn.cursor()
        if tabela == "tabela_107":
            colunas = [
                "ptaCod", "logNsuPta", "logNsuHst", "oprMat", "supMat", "terNum", "ptaCodDsn", "ctrCod",
                "logCrpCod", "logDatMov", "logHorMov", "logNsuAss", "logNsuEsr", "logTrnSta", "logNumPrt",
                "logCnaNum004", "logRef", "l107RegTip", "l107IdcSeqFEB", "l107SeqFEB", "logQtdDiaBlq",
                "logTipRec", "logTipRecHst", "logTipRecAB", "logValPcp014", "logValMul014", "logValJur014",
                "logValAtz014", "logValAbt014", "logValDcn014", "logValAcr014", "logValTot014"
            ]
        elif tabela == "tabela_005":
            colunas = [
                "ptaCod", "logNsuPta", "logNsuHst", "oprMat", "supMat", "terNum", "ptaCodDsn", "ctrCod",
                "logCrpCod", "logDatMov", "logHorMov", "logNsuAss", "logNsuEsr", "logTrnSta", "logNumPrt",
                "ctaNum", "l005DatVal", "docNum", "logValLan014", "logChqQtd003", "logQtdDiaBlq", "ageOri",
                "nsuAnt", "cctTip", "l005CpfClt", "l005IdcLanMlt", "ProtocoloTarifa", "TipoConta"
            ]
        placeholders = ",".join(["%s"] * len(colunas))
        query = f"INSERT INTO {tabela} ({', '.join(colunas)}) VALUES ({placeholders})"
        for row in dados:
            row = [tratar_valores(celula) for celula in row]
            cursor.execute(query, row)
        conn.commit()
        st.success(f"✅ {len(dados)} registros inseridos na tabela {tabela}")
    except Exception as e:
        st.error(f"❌ Erro ao inserir na tabela {tabela}: {e}")
        for i, row in enumerate(dados):
            try:
                cursor.execute(query, row)
            except Exception as e:
                st.error(f"❌ Erro na linha {i+1}: {e} - Dados: {row}")
    finally:
        cursor.close()
        conn.close()

# Função para montar o dashboard de Cálculo Total e Média utilizando os CNP da tabela cnp_data (com situacao = '1')
def get_dashboard_data():
    """
    Consulta a tabela cnp_data para obter os dados dos CNPs com situacao = '1'
    e junta com os cálculos:
      - Valor_005: Calculado a partir da tabela_005 (recolhimento líquido - suprimento líquido)
      - Valor_107: Calculado a partir da tabela_107 (fórmula dos malotes)
      - Cheque_009: Fica zerado
      - Média_Mês: Calculado pela fórmula:
            (Valor_005 - Valor_107 - Cheque_009) / 22
    """
    query_cnp = "SELECT cnp, razao_social as Nome FROM cnp_data WHERE situacao = '1';"
    df_cnp = pd.read_sql(query_cnp, engine)
    
    query_005 = """
        SELECT 
            ptaCod AS cnp,
            (
              (SUM(CASE WHEN ctrCod = 25400 AND logTrnSta <> 'E' THEN logValLan014 ELSE 0 END)
               - SUM(CASE WHEN ctrCod = 25400 AND logTrnSta = 'E' THEN logValLan014 ELSE 0 END)
              )
              -
              (SUM(CASE WHEN ctrCod = 25300 AND logTrnSta <> 'E' THEN logValLan014 ELSE 0 END)
               - SUM(CASE WHEN ctrCod = 25300 AND logTrnSta = 'E' THEN logValLan014 ELSE 0 END)
              )
            ) AS Valor_005
        FROM tabela_005
        GROUP BY ptaCod
    """
    df_005 = pd.read_sql(query_005, engine)
    
    query_107 = """
        SELECT 
            logCnaNum004 AS cnp,
            (
              SUM(logValTot014)
              - SUM(CASE WHEN ptaCod = 245 THEN logValTot014 ELSE 0 END)
              - SUM(CASE WHEN logTrnSta = 'E' THEN logValTot014 ELSE 0 END)
            ) AS Valor_107
        FROM tabela_107
        GROUP BY logCnaNum004
    """
    df_107 = pd.read_sql(query_107, engine)
    
    # Merge dos dados: primeiro os CNP ativos, depois os dados de movimentação e malotes
    df_dash = df_cnp.merge(df_005, on="cnp", how="left")
    df_dash = df_dash.merge(df_107, on="cnp", how="left")
    
    df_dash["Valor_005"] = df_dash["Valor_005"].fillna(0)
    df_dash["Valor_107"] = df_dash["Valor_107"].fillna(0)
    df_dash["Cheque_009"] = 0
    
    # Cálculo da Média_Mês:
    df_dash["Média_Mês"] = (df_dash["Valor_005"] - df_dash["Valor_107"] - df_dash["Cheque_009"]) / 22
    
    return df_dash

# Menu de navegação
menu = st.sidebar.radio("Menu", ["Dashboard", "Importar Planilha", "Calcular Total e Média"])

if menu == "Dashboard":
    st.title("Dashboard")
    st.write("Bem-vindo ao Dashboard! Esse espaço será utilizado posteriormente.")

elif menu == "Importar Planilha":
    st.title("Importar Planilha para o Banco de Dados")
    arquivo = st.file_uploader("Escolha a planilha Excel (movimentação)", type=["xlsx"])
    if arquivo is not None:
        st.write("📂 Extraindo dados... Aguarde ⏳")
        dados_107, dados_005 = extrair_dados(arquivo)
        if dados_107:
            inserir_dados("tabela_107", dados_107)
        if dados_005:
            inserir_dados("tabela_005", dados_005)
        st.write("✔️ Processo concluído!")

elif menu == "Calcular Total e Média":
    st.title("Calcular Total e Média")
    st.write("## Dashboard de Cálculo Total e Média")
    df_dashboard = get_dashboard_data()
    st.dataframe(
        df_dashboard.style.format({
            'Valor_005': '{:,.2f}',
            'Valor_107': '{:,.2f}',
            'Cheque_009': '{:,.2f}',
            'Média_Mês': '{:,.2f}'
        }),
        use_container_width=True,
        height=600
    )
